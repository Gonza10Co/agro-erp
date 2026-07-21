"""
Genera el BOM con despiece (material x pieza, consumo por talla) desde el Excel del cliente.

Fuente: hoja CONSUMOSXREFERENCIA de CONTROL_PRODUCCIÓN_AGROINDUSTRIAL.xlsx
        (Drive del cliente; el archivo NO se versiona).

Regla del cliente (Juan Pablo, 2026-07-09): en esa hoja las celdas SOMBREADAS EN GRIS son
consumos con prueba industrial hecha; las BLANCAS traen valores que no son reales. Se cargan
solo los grises; los blancos entran en cero, a la espera de que terminen el ejercicio.

Uso:
    python tools/generar-despiece.py <ruta-al-xlsx>

Escribe (dentro de backend/prisma/data/basarili/, que está fuera de control de versiones):
    bom-curva.csv      referencia,material,pieza,talla,consumo
    bom-fijo.csv       referencia,material,consumoFijo,mermaPct   (sin los materiales despiezados)
    bom-variantes.csv  variante,referencia,accion,materialObjetivo,pieza,materialNuevo,talla,consumo
                       Reglas de override para las variantes elegidas al pedir:
                       - ECONOMICA: diff contra el despiece base de su referencia
                         (SET_CONSUMO si cambia la curva, ADD si la línea es nueva,
                         REMOVE si la línea base desaparece en la variante).
                       - SP (sin puntera): ADD del contrafuerte preformado con su curva
                         (hoy en CERO: bloques blancos, sin prueba industrial). El REMOVE
                         de la puntera es estructural y lo pone el seed.

Requiere: pip install openpyxl
"""
import re
import shutil
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

RAIZ = Path(__file__).resolve().parent.parent
DATA = RAIZ / "backend" / "prisma" / "data" / "basarili"

HOJA = "CONSUMOSXREFERENCIA"
FILA_TALLAS = 3
COL_INI = 3  # columna C
GRISES = {"FFCCCCCC", "FFB7B7B7"}

# La 106 = RESORTADA quedó confirmada el 2026-07-09. Las variantes ECONOMICA y S/P
# (elegidas al pedir) ya NO se descartan: salen como reglas en bom-variantes.csv.
REFS_EXCLUIDAS: set[str] = set()

# Sufijo del nombre -> código de la pieza en el catálogo `Pieza`.
# El orden importa: "SOPORTE LATERAL" debe probarse antes que "LATERAL".
PIEZAS = [
    ("SOPORTE LATERAL", "SOPORTE_LATERAL"),
    ("CAPELLADA", "CAPELLADA"),
    ("EMPEINE", "EMPEINE"),
    ("LATERALES", "LATERAL"),
    ("LATERAL", "LATERAL"),
    ("TALON", "TALON"),
    ("BOTELLA", "BOTELLA"),
    ("CANA", "CANA"),
    ("CAÑA", "CANA"),
    ("FUELLE", "FUELLE"),
    ("CUELLO", "CUELLO"),
    ("LENGUA", "LENGUA"),
    ("VISTA", "VISTA"),
    ("CHAPETA", "CHAPETA"),
    ("CORDONERA", "CORDONERA"),
    ("PLANTILLA", "PLANTILLA"),
    ("ECONOMIZADOR", "ECONOMIZADOR"),
]

# Nombre del material en la hoja (ya sin la pieza) -> código en materiales.csv.
# Las notas marcan dónde se eligió un representante para un nombre genérico.
MATERIALES = {
    "MICROPIEL NEGRA/CAFE": "PMIC191",  # genérico negra/café -> NEGRA LISA
    "MICROPIEL GRABADA": "PMIC190",
    "MICROFIBRA SINTETIC PVC": "PMIC187",
    "LAZZIO": "PLAZ127",  # LAZZIO/ROSSY
    "SINTETIC PVC": "PSIN262",
    "MALLA FORRO": "PMAL278",  # genérico -> MALLA FALCAO NEGRO
    "STROBEL": "PSTR263",  # STROBEL 300 GRMS
    "VAMPLING FORRO": "PVAM265",  # VAMPLING LININF
    "CONTRAFUERTE COUNTER DOUBLE SID": "PCON44",
    # El preformado S/P es un material distinto (reemplaza puntera + contrafuerte en la
    # variante "sin puntera"); antes colisionaba con PCON44 y se descartaba.
    "CONTRAFUERTE COUNTER DOUBLE SID S/P": "PCON44SP",
    "LAMBRILLA 2,5": "PLAM126",
    "LAMBRILLA EVA 2,5 BULLON": "PLAM126",  # cliente: BULLON EVA = LAMBRILLA 2,5
    "LAMBRILLLA EVA 2,5 BULLON": "PLAM126",  # sic: triple L, typo del Excel en la ref 106
    "PRODUPIQUETH 2,5": "PPRO251",
    "ESPUMA 1,2MM BULLON": "PESP78",
    "YUMBOLON #05 BULLON": "PYUM269",
    "YUMBOLON #08 BULLON": "PYUM270",
}


# Consumos del bom-fijo histórico que estaban en la unidad FÍSICA (metros, unidades)
# cuando el material se compra y costea por EMPAQUE (cono, gruesa). Sin esta conversión
# el costeo multiplica metros × precio-del-cono y el costo del par sale en millones.
#   - Hilos ref 101: venían en metros; las refs 102-105 ya estaban en conos.
#     0.0086 = 43.02 m ÷ 5000 m/cono · 0.0036 = 27.16 m ÷ 7500 m/cono.
#   - Cordón: la GRUESA trae 144 PARES de cordones (confirmado por Juan Pablo,
#     2026-07-10). Un par de botas consume un par de cordones → 1/144 = 0.0069.
FIJOS_CORREGIDOS: dict[tuple[str, str], float] = {
    ("101", "PHIL111"): 0.0086,
    ("101", "PHIL112"): 0.0036,
    ("101", "PCOR52"): 0.0069,
    ("102", "PCOR52"): 0.0069,
    ("103", "PCOR52"): 0.0069,
    ("104", "PCOR52"): 0.0069,
}


def es_gris(celda) -> bool:
    f = celda.fill
    if f is None or f.patternType is None:
        return False
    try:
        return str(f.fgColor.rgb) in GRISES
    except Exception:
        return False


def partir_pieza(nombre: str) -> tuple[str, str | None]:
    """Separa 'MICROPIEL NEGRA/CAFE CAPELLADA' en ('MICROPIEL NEGRA/CAFE', 'CAPELLADA')."""
    for sufijo, codigo in PIEZAS:
        if nombre.endswith(sufijo):
            return nombre[: -len(sufijo)].strip(), codigo
    return nombre, None


def main() -> int:
    if len(sys.argv) < 2:
        print(__doc__)
        return 2
    xlsx = Path(sys.argv[1])
    if not xlsx.exists():
        print(f"No existe el archivo: {xlsx}")
        return 1

    ws = openpyxl.load_workbook(xlsx, data_only=True)[HOJA]
    tallas = [
        ws.cell(FILA_TALLAS, c).value
        for c in range(COL_INI, ws.max_column + 1)
    ]
    tallas = [int(t) for t in tallas if isinstance(t, (int, float))]

    # clave (ref, material, pieza) -> (validado, {talla: consumo}, nombre en la hoja)
    lineas: dict[tuple[str, str, str], tuple[bool, dict[int, float], str]] = {}
    # clave (variante, ref, material, pieza) -> (validado, {talla: consumo}, nombre)
    var_lineas: dict[tuple[str, str, str, str], tuple[bool, dict[int, float], str]] = {}
    despiezados: dict[str, set[str]] = defaultdict(set)
    sin_mapeo: set[str] = set()
    colisiones: list[str] = []
    combos_omitidos: set[str] = set()
    validados = sin_validar = 0

    for r in range(1, ws.max_row + 1):
        a, b = ws.cell(r, 1).value, ws.cell(r, 2).value
        if not isinstance(a, str) or not isinstance(b, str):
            continue
        if b.strip().upper() in ("PARES", "CONSUMO", "MATERIA PRIMA"):
            continue
        m = re.match(r"^(\d{3})\s+(.*)$", a.strip())
        if not m:
            continue
        ref, nombre = m.group(1), m.group(2).strip().upper()
        if ref in REFS_EXCLUIDAS:
            continue

        variante: str | None = None
        nombre_bom = nombre
        if nombre.startswith("ECONOMICA"):
            variante = "ECONOMICA"
            nombre_bom = nombre[len("ECONOMICA"):].strip()

        base, pieza = partir_pieza(nombre_bom)
        codigo = MATERIALES.get(base)
        if not codigo:
            sin_mapeo.add(f"{ref}: {base}")
            continue
        if codigo == "PCON44SP":
            if variante == "ECONOMICA":
                # Combo ECONOMICA+S/P: se compone solo (reglas de ambas opciones); no
                # necesita fila propia mientras el bloque siga en blanco.
                combos_omitidos.add(f"{ref}: {nombre}")
                continue
            variante = "SP"

        # El color manda: un bloque sin gris no tiene prueba industrial.
        bloque_validado = any(es_gris(ws.cell(r, c)) for c in range(COL_INI, COL_INI + len(tallas)))
        if variante is None:
            if bloque_validado:
                validados += 1
            else:
                sin_validar += 1

        curva = {}
        for i, talla in enumerate(tallas):
            celda = ws.cell(r, COL_INI + i)
            valor = celda.value if isinstance(celda.value, (int, float)) else 0
            curva[talla] = round(float(valor), 6) if (bloque_validado and es_gris(celda)) else 0

        if variante is not None:
            var_lineas[(variante, ref, codigo, pieza or "")] = (bloque_validado, curva, nombre)
            continue

        despiezados[ref].add(codigo)

        # Dos nombres de la hoja pueden caer en el mismo (material, pieza). Gana el que
        # tenga prueba industrial; si chocaran dos validados, hay que desambiguar el catálogo.
        clave = (ref, codigo, pieza or "")
        previo = lineas.get(clave)
        if previo:
            antes_validado, _, antes_nombre = previo
            colisiones.append(
                f"{ref}: «{antes_nombre}» y «{nombre}» comparten {codigo}"
                f"{'/' + pieza if pieza else ' (bota completa)'}"
                + (" — AMBOS VALIDADOS" if antes_validado and bloque_validado else "")
            )
            if antes_validado and not bloque_validado:
                continue  # el validado ya está puesto
        lineas[clave] = (bloque_validado, curva, nombre)

    if sin_mapeo:
        print("MATERIALES SIN MAPEO (no se cargan):")
        for s in sorted(sin_mapeo):
            print(f"  - {s}")
    if colisiones:
        print("COLISIONES de (material, pieza) — gana el validado:")
        for c in sorted(set(colisiones)):
            print(f"  - {c}")

    filas = [
        (ref, cod, pz, talla, cons)
        for (ref, cod, pz), (_, curva, _n) in sorted(lineas.items())
        for talla, cons in sorted(curva.items())
    ]

    # --- bom-curva.csv
    curva = DATA / "bom-curva.csv"
    with curva.open("w", encoding="utf-8", newline="") as f:
        f.write("referencia,material,pieza,talla,consumo\n")
        for ref, cod, pz, talla, cons in filas:
            f.write(f"{ref},{cod},{pz},{talla},{cons}\n")

    # --- bom-fijo.csv: quitar los materiales que ahora vienen despiezados con curva,
    # o la receta tendría el mismo material dos veces (una plana y otra por pieza).
    fijo = DATA / "bom-fijo.csv"
    respaldo = fijo.with_suffix(".csv.pre-despiece")
    if fijo.exists():
        # Idempotente: la fuente es siempre el respaldo original, nunca el ya filtrado.
        if not respaldo.exists():
            shutil.copy(fijo, respaldo)
        crudo = respaldo.read_text(encoding="utf-8").strip().splitlines()
        header, cuerpo = crudo[0], crudo[1:]
        conservadas = []
        corregidas = 0
        for l in cuerpo:
            p = l.split(",")
            if p[1] in despiezados.get(p[0], set()):
                continue
            nuevo = FIJOS_CORREGIDOS.get((p[0], p[1]))
            if nuevo is not None and float(p[2]) != nuevo:
                l = f"{p[0]},{p[1]},{nuevo},{p[3] if len(p) > 3 else ''}"
                corregidas += 1
            conservadas.append(l)
        fijo.write_text("\n".join([header, *conservadas]) + "\n", encoding="utf-8")
        quitadas = len(cuerpo) - len(conservadas)
        print(
            f"bom-fijo.csv: {len(cuerpo)} líneas originales -> {len(conservadas)} conservadas"
            f" · {corregidas} consumos convertidos a unidad de empaque"
        )
    else:
        quitadas = 0

    # --- bom-variantes.csv: reglas de override por (material, pieza) para las variantes.
    # (variante, ref, accion, materialObjetivo, pieza, materialNuevo, curva|None)
    reglas: list[tuple[str, str, str, str, str, str, dict[int, float] | None]] = []

    # ECONOMICA: diff contra el despiece base de su misma referencia.
    refs_econ = sorted({ref for (v, ref, _c, _p) in var_lineas if v == "ECONOMICA"})
    for ref in refs_econ:
        base_ref = {
            (cod, pz): crv for (r, cod, pz), (_v, crv, _n) in lineas.items() if r == ref
        }
        econ_ref = {
            (cod, pz): crv
            for (v, r, cod, pz), (_va, crv, _n) in var_lineas.items()
            if v == "ECONOMICA" and r == ref
        }
        for (cod, pz), crv in sorted(econ_ref.items()):
            if (cod, pz) in base_ref:
                if base_ref[(cod, pz)] != crv:
                    reglas.append(("ECONOMICA", ref, "SET_CONSUMO", cod, pz, "", crv))
            else:
                reglas.append(("ECONOMICA", ref, "ADD", "", pz, cod, crv))
        for (cod, pz) in sorted(base_ref):
            if (cod, pz) not in econ_ref:
                reglas.append(("ECONOMICA", ref, "REMOVE", cod, pz, "", None))

    # S/P: ADD del contrafuerte preformado con su curva (hoy en cero, bloques blancos).
    # El REMOVE de la puntera es estructural: lo resuelve el seed contra el BOM de la ref.
    for (v, ref, cod, pz), (_va, crv, _n) in sorted(var_lineas.items()):
        if v == "SP":
            reglas.append(("SP", ref, "ADD", "", pz, cod, crv))

    var_csv = DATA / "bom-variantes.csv"
    with var_csv.open("w", encoding="utf-8", newline="") as f:
        f.write("variante,referencia,accion,materialObjetivo,pieza,materialNuevo,talla,consumo\n")
        for v, ref, acc, obj, pz, nuevo, crv in reglas:
            if crv is None:
                f.write(f"{v},{ref},{acc},{obj},{pz},{nuevo},,\n")
            else:
                for talla, cons in sorted(crv.items()):
                    f.write(f"{v},{ref},{acc},{obj},{pz},{nuevo},{talla},{cons}\n")

    if combos_omitidos:
        print("COMBOS variante omitidos (se componen solos):")
        for c in sorted(combos_omitidos):
            print(f"  - {c}")

    refs = sorted(despiezados)
    resumen_var = defaultdict(lambda: defaultdict(int))
    for v, ref, acc, *_ in reglas:
        resumen_var[v][acc] += 1
    print(
        f"Despiece: {validados} bloques validados (gris) · {sin_validar} en cero (blanco) · "
        f"{len(filas)} filas de curva · refs {', '.join(refs)}"
    )
    for v, accs in sorted(resumen_var.items()):
        detalle = " · ".join(f"{a}×{n}" for a, n in sorted(accs.items()))
        print(f"Variante {v}: {detalle}")
    print(f"bom-fijo.csv: {quitadas} líneas retiradas (esos materiales ahora van por pieza)")
    print(f"Escrito: {curva.name}, {fijo.name} y {var_csv.name} (respaldo en bom-fijo.csv.pre-despiece)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
